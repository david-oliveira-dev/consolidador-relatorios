"""Geração do relatório final em Excel (.xlsx).

Produz um arquivo com três partes:
    - aba "Dados": todos os registros consolidados e limpos
    - aba "Resumo": tabelas de totais por produto e por mês
    - um gráfico de barras (faturamento por produto) na aba "Resumo"

Decisão de design: as tabelas são escritas com pandas e a formatação
(cabeçalho, larguras, formato de moeda, gráfico) é feita com openpyxl, que é a
engine que o pandas já usa para .xlsx — então não adicionamos dependência nova.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

# Formatos numéricos no padrão de exibição do Excel.
FORMATO_MOEDA = '"R$" #,##0.00'
FORMATO_INTEIRO = "#,##0"
FORMATO_DATA = "yyyy-mm-dd"

# Estilo do cabeçalho das tabelas.
_FONTE_CABECALHO = Font(bold=True, color="FFFFFF")
_FUNDO_CABECALHO = PatternFill("solid", fgColor="305496")  # azul escuro
_CENTRO = Alignment(horizontal="center")


# ---------------------------------------------------------------------------
# Cálculo dos resumos
# ---------------------------------------------------------------------------
def resumo_por_produto(df: pd.DataFrame) -> pd.DataFrame:
    """Total, ticket médio, quantidade e nº de vendas por produto (desc)."""
    resumo = (
        df.groupby("produto")
        .agg(
            faturamento=("valor", "sum"),
            ticket_medio=("valor", "mean"),
            quantidade_total=("quantidade", "sum"),
            num_vendas=("valor", "count"),
        )
        .reset_index()
        .sort_values("faturamento", ascending=False)
    )
    return resumo


def resumo_por_mes(df: pd.DataFrame) -> pd.DataFrame:
    """Faturamento e nº de vendas por mês (ignora linhas sem data válida)."""
    com_data = df.dropna(subset=["data"]).copy()
    # to_period('M') agrupa por mês; viramos texto "2025-01" para exibir bonito.
    com_data["mes"] = com_data["data"].dt.to_period("M").astype(str)
    resumo = (
        com_data.groupby("mes")
        .agg(faturamento=("valor", "sum"), num_vendas=("valor", "count"))
        .reset_index()
        .sort_values("mes")
    )
    return resumo


# ---------------------------------------------------------------------------
# Helpers de formatação (openpyxl)
# ---------------------------------------------------------------------------
def _estilizar_cabecalho(ws: Worksheet, linha: int, n_colunas: int) -> None:
    """Aplica negrito, cor e centralização à linha de cabeçalho informada."""
    for col in range(1, n_colunas + 1):
        celula = ws.cell(row=linha, column=col)
        celula.font = _FONTE_CABECALHO
        celula.fill = _FUNDO_CABECALHO
        celula.alignment = _CENTRO


def _aplicar_formato_numerico(
    ws: Worksheet, header_row: int, n_linhas: int, formatos: dict[str, str]
) -> None:
    """Aplica formato (moeda/inteiro/data) às colunas pelo nome do cabeçalho."""
    cabecalhos = {
        ws.cell(row=header_row, column=c).value: c
        for c in range(1, ws.max_column + 1)
    }
    for nome, fmt in formatos.items():
        col = cabecalhos.get(nome)
        if col is None:
            continue
        for linha in range(header_row + 1, header_row + 1 + n_linhas):
            ws.cell(row=linha, column=col).number_format = fmt


def _ajustar_larguras(ws: Worksheet) -> None:
    """Ajusta a largura de cada coluna ao maior conteúdo (com folga)."""
    for coluna in ws.columns:
        # Ignora células de células mescladas/sem coordenada de coluna.
        letras = [c.column_letter for c in coluna if hasattr(c, "column_letter")]
        if not letras:
            continue
        largura = max((len(str(c.value)) for c in coluna if c.value is not None), default=8)
        ws.column_dimensions[letras[0]].width = min(largura + 2, 40)


# ---------------------------------------------------------------------------
# Orquestrador
# ---------------------------------------------------------------------------
def gerar_relatorio_excel(df: pd.DataFrame, caminho_saida: Path) -> Path:
    """Escreve o relatório completo em `caminho_saida` e retorna o caminho.

    A pasta de saída é criada se não existir.
    """
    caminho_saida.parent.mkdir(parents=True, exist_ok=True)

    rp = resumo_por_produto(df)
    rm = resumo_por_mes(df)

    # Posições (1-indexado) das tabelas na aba Resumo.
    linha_titulo_prod = 1
    header_prod = 2  # cabeçalho da tabela de produtos
    # tabela de meses começa alguns espaços abaixo da de produtos
    linha_titulo_mes = header_prod + len(rp) + 3
    header_mes = linha_titulo_mes + 1

    with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
        # Aba Dados
        df.to_excel(writer, sheet_name="Dados", index=False)
        # Aba Resumo: duas tabelas (startrow é 0-indexado no pandas)
        rp.to_excel(writer, sheet_name="Resumo", index=False, startrow=header_prod - 1)
        rm.to_excel(writer, sheet_name="Resumo", index=False, startrow=header_mes - 1)

        ws_dados = writer.sheets["Dados"]
        ws_resumo = writer.sheets["Resumo"]

        # --- Aba Dados: cabeçalho + formatos + larguras ---
        _estilizar_cabecalho(ws_dados, 1, ws_dados.max_column)
        _aplicar_formato_numerico(
            ws_dados,
            header_row=1,
            n_linhas=len(df),
            formatos={
                "data": FORMATO_DATA,
                "valor": FORMATO_MOEDA,
                "quantidade": FORMATO_INTEIRO,
            },
        )
        _ajustar_larguras(ws_dados)

        # --- Aba Resumo: títulos das tabelas ---
        ws_resumo.cell(row=linha_titulo_prod, column=1, value="Faturamento por Produto").font = Font(bold=True, size=12)
        ws_resumo.cell(row=linha_titulo_mes, column=1, value="Faturamento por Mês").font = Font(bold=True, size=12)

        # cabeçalhos das duas tabelas
        _estilizar_cabecalho(ws_resumo, header_prod, len(rp.columns))
        _estilizar_cabecalho(ws_resumo, header_mes, len(rm.columns))

        # formatos da tabela de produtos
        _aplicar_formato_numerico(
            ws_resumo, header_prod, len(rp),
            {"faturamento": FORMATO_MOEDA, "ticket_medio": FORMATO_MOEDA,
             "quantidade_total": FORMATO_INTEIRO, "num_vendas": FORMATO_INTEIRO},
        )
        # formatos da tabela de meses
        _aplicar_formato_numerico(
            ws_resumo, header_mes, len(rm),
            {"faturamento": FORMATO_MOEDA, "num_vendas": FORMATO_INTEIRO},
        )
        _ajustar_larguras(ws_resumo)

        # --- Gráfico de barras: faturamento por produto ---
        grafico = BarChart()
        grafico.title = "Faturamento por Produto"
        grafico.y_axis.title = "R$"
        grafico.x_axis.title = "Produto"
        grafico.legend = None

        # coluna 2 (faturamento) é a série; coluna 1 (produto) são as categorias.
        dados_ref = Reference(
            ws_resumo, min_col=2, min_row=header_prod, max_row=header_prod + len(rp)
        )
        cats_ref = Reference(
            ws_resumo, min_col=1, min_row=header_prod + 1, max_row=header_prod + len(rp)
        )
        grafico.add_data(dados_ref, titles_from_data=True)
        grafico.set_categories(cats_ref)
        grafico.height = 8
        grafico.width = 16
        # ancora o gráfico à direita das tabelas
        ws_resumo.add_chart(grafico, "H2")

    return caminho_saida
